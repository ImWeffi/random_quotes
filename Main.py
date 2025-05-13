import requests
from datetime import datetime
from bs4 import BeautifulSoup
import random
import os
from openpyxl import Workbook, load_workbook

top_tags = [
    "love", "inspirational", "life", "humor", "books",
    "reading", "friendship", "friends", "truth", "simile"
]

def paradi_darbibas():
    darbiba = ["skatīt citātu", "saglabāt citātu", "cita tēma", "iziet"]
    print("\nizvēlies darbību:")
    for id, nosaukums in enumerate(darbiba, 1):
        print(f"{id}. {nosaukums}")
    return darbiba

def paradi_temas():
    print("\npieejamās tēmas:")
    for idx, tag in enumerate(top_tags, 1):
        print(f"{idx}. {tag}")

def iegut_citatus(tag):
    adrese = f"https://quotes.toscrape.com/tag/{tag}/"
    lapa = requests.get(adrese)
    if lapa.status_code == 200:
        lapas_saturs = BeautifulSoup(lapa.content, "html.parser")
        citatu_bloki = lapas_saturs.find_all(class_="quote")
        citatas = []
        for bloks in citatu_bloki:
            teksts = bloks.find(class_="text").get_text()
            autors = bloks.find(class_="author").get_text()
            citatas.append((teksts, autors))
        return citatas
    else:
        print("neizdevās ielādēt lapu.")
        return []

def saglabat_excel( autors, citats, tema):
    fails = "Citati.xlsx"
    datums = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    if os.path.exists(fails):
        wb = load_workbook(fails)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append([ "Autors", "Citāts", "Tēma", "Datums"])

    dublikats = False
    for rinda in ws.iter_rows(min_row=2, values_only=True):
        if rinda[0] == autors and rinda[1] == citats and rinda[2] == tema:
            dublikats = True
            break

    if dublikats:
        print("šis citāts jau ir saglabāts.")
    else:
        ws.append([ autors, citats, tema, datums])
        wb.save(fails)
        print("citāts saglabāts Excel failā.")

def izveleties_temu():
    paradi_temas()
    izvele = input("ievadi tēmas numuru (1–10): ")
    try:
        izvele = int(izvele)
        if 1 <= izvele <= len(top_tags):
            return top_tags[izvele - 1]
        else:
            print("nepareizs numurs.")
            return None
    except ValueError:
        print("ievadi skaitli.")
        return None

izveletais_tag = None
while izveletais_tag is None:
    izveletais_tag = izveleties_temu()

visi_citati = iegut_citatus(izveletais_tag)
if not visi_citati:
    print("nav citātu šai tēmai.")
    exit()

izveletais_citats, autors = random.choice(visi_citati)
print(f"\n💬 {izveletais_citats}\n— {autors}")

while True:
    darbibas = paradi_darbibas()
    darbiba_izvele = input("ievadi darbības numuru: ")

    if darbiba_izvele == "1":
        izveletais_citats, autors = random.choice(visi_citati)
        print(f"\n💬 {izveletais_citats}\n— {autors}")
    elif darbiba_izvele == "2":
        saglabat_excel( autors, izveletais_citats, izveletais_tag)
    elif darbiba_izvele == "3":
        izveletais_tag = None
        while izveletais_tag is None:
            izveletais_tag = izveleties_temu()
        visi_citati = iegut_citatus(izveletais_tag)
        if not visi_citati:
            print("nav citātu šai tēmai.")
            continue
        izveletais_citats, autors = random.choice(visi_citati)
        print(f"\n💬 {izveletais_citats}\n— {autors}")
    elif darbiba_izvele == "4":
        print("programma beidzas.")
        break
    else:
        print("nepareiza izvēle.")